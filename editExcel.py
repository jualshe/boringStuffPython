import openpyxl
import os

workbook = openpyxl.Workbook()
print(workbook)

print(workbook.get_sheet_names())

sheet = workbook.get_sheet_by_name('Sheet')
print(sheet)

print(sheet['A1'].value)

sheet['A1'] = 42
sheet['A2'] = 'Hello'

os.chdir('/Users/julia/Documents')

sheet2 = workbook.create_sheet()
print(sheet2.title)
sheet2.title = 'NEW SHEET TITLE'
workbook.create_sheet(index=0, title='My other sheet')
print(workbook.get_sheet_names())

workbook.save('example3.xlsx')
