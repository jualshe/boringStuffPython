import openpyxl
import os

os.chdir('/Users/julia/dev')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(workbook.get_sheet_names())
print(type(sheet))

cell = sheet['A1']
print(str(cell.value))

print(str(sheet['A1'].value))
print(str(sheet['C1'].value))

# these are the same
print(sheet['B1'].value)
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
